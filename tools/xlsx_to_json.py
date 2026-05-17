# -*- coding: utf-8 -*-
"""Liest content_template.xlsx und schreibt assets/content/it/ neu.

Output:
  assets/content/it/manifest.json
  assets/content/it/lessons/lesson.*.json

contentVersion wird automatisch hochgezaehlt (alter Wert + 1), damit der
Seeder beim naechsten App-Start die DB re-seedet.
"""
import argparse
import json
import os
import re
import sys
from collections import OrderedDict
from openpyxl import load_workbook

REPO_ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_XLSX  = os.path.join(REPO_ROOT, "tools", "content_template.xlsx")
ASSETS_DIR    = os.path.join(REPO_ROOT, "assets", "content", "it")
LESSONS_DIR   = os.path.join(ASSETS_DIR, "lessons")
MANIFEST_PATH = os.path.join(ASSETS_DIR, "manifest.json")

VOCAB_HEADERS = [
    "slug", "italiano", "deutsch", "wortart", "artikel", "plural",
    "beispielsatz_it", "beispielsatz_de", "notiz", "suggested",
]
SENTENCE_HEADERS = ["italiano", "deutsch", "notiz"]
VALID_SUGGESTED = {"", "mc", "pair", "typing"}


def _cell(row, idx):
    if idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def parse_manifest_sheet(sheet):
    """Liest CHAPTERS- und LESSONS-Tabelle aus dem _manifest-Sheet."""
    rows = list(sheet.iter_rows(values_only=True))
    chapters, lessons = [], []
    section = None
    header_skip = 0
    for r in rows:
        first = _cell(r, 0)
        if first == "CHAPTERS":
            section = "chapters"; header_skip = 1; continue
        if first == "LESSONS":
            section = "lessons"; header_skip = 1; continue
        if header_skip > 0:
            header_skip -= 1; continue
        if not first:
            section = None; continue
        if section == "chapters":
            chapters.append({
                "id": _cell(r, 0),
                "title": _cell(r, 1),
                "sortOrder": int(_cell(r, 2) or 0),
            })
        elif section == "lessons":
            lessons.append({
                "sheet_name": _cell(r, 0),
                "lesson_id": _cell(r, 1),
                "chapter_id": _cell(r, 2),
                "title": _cell(r, 3),
                "sort_order": int(_cell(r, 4) or 0),
            })
    return chapters, lessons


def parse_lesson_sheet(sheet):
    """Liest VOKABELN- und SAETZE-Tabelle aus einem Lesson-Sheet."""
    rows = list(sheet.iter_rows(values_only=True))
    vocab, sentences = [], []
    section = None
    header_skip = 0
    for r in rows:
        first = _cell(r, 0)
        if first == "VOKABELN":
            section = "vocab"; header_skip = 2; continue  # header + hint
        if first in ("SÄTZE", "SAETZE"):
            section = "sentences"; header_skip = 2; continue
        if header_skip > 0:
            header_skip -= 1; continue
        if not first:
            section = None; continue
        if section == "vocab":
            vocab.append({h: _cell(r, i) for i, h in enumerate(VOCAB_HEADERS)})
        elif section == "sentences":
            sentences.append({h: _cell(r, i) for i, h in enumerate(SENTENCE_HEADERS)})
    return vocab, sentences


def vocab_to_step(v):
    if not v["slug"] or not v["italiano"] or not v["deutsch"]:
        raise ValueError(f"Pflichtfeld fehlt in Vokabel: {v}")
    step = OrderedDict()
    step["kind"] = "vocab"
    step["itemId"] = v["slug"]
    step["target"] = v["italiano"]
    step["native"] = v["deutsch"]
    if v["wortart"]:
        step["partOfSpeech"] = v["wortart"]
    suggested = v["suggested"].lower()
    if suggested not in VALID_SUGGESTED:
        raise ValueError(f"Ungueltiges 'suggested' fuer {v['slug']}: {suggested!r}")
    if suggested and suggested != "mc":
        step["suggested"] = suggested
    return step


def sentence_to_step(s, exercise_id):
    if not s["italiano"] or not s["deutsch"]:
        raise ValueError(f"Pflichtfeld fehlt in Satz: {s}")
    step = OrderedDict()
    step["kind"] = "sentence_builder"
    step["exerciseId"] = exercise_id
    step["nativePrompt"] = s["deutsch"]
    step["targetWords"] = s["italiano"].split(" ")
    return step


def lesson_suffix(lesson_id):
    """'lesson.01-01' -> '01-01'."""
    m = re.match(r"^lesson\.(.+)$", lesson_id)
    if not m:
        raise ValueError(f"Unerwartete lesson_id: {lesson_id!r}")
    return m.group(1)


def bump_content_version():
    if not os.path.exists(MANIFEST_PATH):
        return 1
    with open(MANIFEST_PATH, encoding="utf-8") as f:
        return int(json.load(f).get("contentVersion", 0)) + 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    if "_manifest" not in wb.sheetnames:
        print("ERROR: _manifest-Sheet fehlt", file=sys.stderr); sys.exit(1)

    chapters, lessons = parse_manifest_sheet(wb["_manifest"])

    # Sanity: lesson_ids eindeutig, chapter_ids existieren
    chapter_ids = {c["id"] for c in chapters}
    seen_lids, seen_slugs = set(), {}
    lesson_payloads = []

    for l in lessons:
        sn = l["sheet_name"]
        if sn not in wb.sheetnames:
            raise ValueError(f"Sheet '{sn}' fehlt im xlsx")
        if l["lesson_id"] in seen_lids:
            raise ValueError(f"Duplikat lesson_id: {l['lesson_id']}")
        if l["chapter_id"] not in chapter_ids:
            raise ValueError(f"Unbekanntes chapter_id: {l['chapter_id']} in {l['lesson_id']}")
        seen_lids.add(l["lesson_id"])

        vocab, sentences = parse_lesson_sheet(wb[sn])
        steps = []
        for v in vocab:
            step = vocab_to_step(v)
            slug = step["itemId"]
            if slug in seen_slugs and seen_slugs[slug] != l["lesson_id"]:
                raise ValueError(
                    f"Slug {slug!r} wird mehrfach genutzt: "
                    f"{seen_slugs[slug]} und {l['lesson_id']}")
            seen_slugs[slug] = l["lesson_id"]
            steps.append(step)
        suffix = lesson_suffix(l["lesson_id"])
        for i, s in enumerate(sentences, start=1):
            ex_id = f"ex.{suffix}.s{i:02d}"
            steps.append(sentence_to_step(s, ex_id))

        lesson_payloads.append({
            "id": l["lesson_id"],
            "chapterId": l["chapter_id"],
            "sortOrder": l["sort_order"],
            "title": l["title"],
            "steps": steps,
        })

    # Manifest aufbauen
    lessons_by_chapter = {c["id"]: [] for c in chapters}
    for l in lessons:
        lessons_by_chapter[l["chapter_id"]].append(
            (l["sort_order"], f"lessons/lesson.{lesson_suffix(l['lesson_id'])}.json"))
    for ch_id in lessons_by_chapter:
        lessons_by_chapter[ch_id].sort()

    manifest = OrderedDict()
    manifest["schemaVersion"] = 1
    manifest["contentVersion"] = bump_content_version()
    manifest["targetLang"] = "it"
    manifest["nativeLang"] = "de"
    manifest["chapters"] = [
        OrderedDict([
            ("id", c["id"]),
            ("title", c["title"]),
            ("sortOrder", c["sortOrder"]),
            ("lessonFiles", [p for _, p in lessons_by_chapter[c["id"]]]),
        ])
        for c in sorted(chapters, key=lambda c: c["sortOrder"])
    ]

    print(f"Chapters: {len(chapters)}  Lessons: {len(lesson_payloads)}  "
          f"Vocab: {sum(1 for lp in lesson_payloads for s in lp['steps'] if s['kind']=='vocab')}  "
          f"Sentences: {sum(1 for lp in lesson_payloads for s in lp['steps'] if s['kind']=='sentence_builder')}")
    print(f"contentVersion: {manifest['contentVersion']}")

    if args.dry_run:
        print("Dry-run — nichts geschrieben.")
        return

    os.makedirs(LESSONS_DIR, exist_ok=True)
    # Alte Lesson-Files weg
    for f in os.listdir(LESSONS_DIR):
        if f.startswith("lesson.") and f.endswith(".json"):
            os.remove(os.path.join(LESSONS_DIR, f))

    for lp in lesson_payloads:
        path = os.path.join(LESSONS_DIR, f"lesson.{lesson_suffix(lp['id'])}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(lp, f, ensure_ascii=False, indent=2)
            f.write("\n")

    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"OK -> {ASSETS_DIR}")


if __name__ == "__main__":
    main()
